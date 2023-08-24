import openpyxl as transact
from openpyxl.chart import BarChart, Reference

wb = transact.load_workbook('transact.xlsx')
sheet = wb['Sheet1']
cell = sheet['a2']
cell = sheet.cell[1, 2]

for row in range(2, sheet.max_row + 1): # diese loop  
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

val = Reference (sheet, min_row=2,max_row=sheet.max_row,min_col = 4, max_cpl = 4 )

chart = BarChart()
chart.add_data(val)
sheet.add_chart(chart, 'g5')


wb.save('transact2.xlsx')

